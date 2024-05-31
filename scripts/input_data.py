from openpyxl import load_workbook
from openpyxl import Workbook

def get_workbook(path):
    return load_workbook(path)

def get_worksheet(workbook):
    return workbook['GLDAS_INPUT']

def get_worksheet_by_name(workbook, name):
    return workbook[name]

def get_folder(row, worksheet):
    return '{}'.format(worksheet['B{}'.format(row)].value)

def get_data_date(row, worksheet):
    return '{}'.format(worksheet['C{}'.format(row)].value)

def save_workbook(path):
    workbook = Workbook()
    workbook.save(path)